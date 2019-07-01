from bs4 import BeautifulSoup
import openpyxl
import re

LOADFILE = './data.html'
SAVEFILE = './jira.xlsx'

TAG_RE = re.compile(r'<[^>]+>')

def main():
    # File read
    with open(LOADFILE, 'r', encoding='utf8') as f:
        raw_data = f.read()
    
    # Set soup
    soup = BeautifulSoup(raw_data, 'html.parser')
    
    # Get data
    issue       = [data.text for data in soup.select('a.issue-link')]
    summary     = [data.text.strip() for data in soup.select('td.summary')]             # strip
    description = [remove_tags(data.text).strip() for data in soup.select('td.description')]    # Remove html tags
    
    # Insert to data
    wb = openpyxl.Workbook()
    ws = wb.active

    for number in range(len(issue)):
        ws['A' + str(number+1)] = issue[number]
        ws['B' + str(number+1)] = summary[number]
        ws['C' + str(number+1)] = description[number]

    
    # Save the file
    wb.save(SAVEFILE)

def remove_tags(text):
    text = text.replace('<br>','\n').replace('&nbsp;',' ').replace('&gt;', '>')
    return TAG_RE.sub('', text)

if __name__ == '__main__':
    main()