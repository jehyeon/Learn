# -*- coding: utf-8 -*-
# FileName  : inserted.py
# Author    : Lee, Jehyeon
# Date      : 20180607

# Importing packages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

import os
import re
import datetime
import configparser
import make_pure_data as pd
import pt_json as pj


'''
    Use config init
'''

# DEFAULT config file name
CONFIG_FILE = 'config.ini'
SAVE_ROUTE = 'inserted_data'

config = configparser.ConfigParser()

# Make config.ini
def makeConfig():
    # config = configparser.ConfigParser()
    config['DEFAULT']['file_route'] = './can-central/primary/'
    config['DEFAULT']['save_route'] = './' + SAVE_ROUTE + '/'
    config['DEFAULT']['full_goal_name'] = 'True'
    # config['DEFAULT']['full_parameter_name'] = 'True'
    config['DEFAULT']['full_utterance'] = 'False'
    config['DEFAULT']['classify_dictionary'] = 'False'
    config['DEFAULT']['pattern_view'] = 'False'

    with open(CONFIG_FILE, 'w') as configfile:
        config.write(configfile)
    print("Make config.ini is done")

    return config

# Read config.ini
def readConfig():
    # config = configparser.ConfigParser()
    config.read(CONFIG_FILE)
    if not 'app_name' in config['DEFAULT']:
        print("Insert your app name: ", end="")
        config['DEFAULT']['app_name'] = input()
        with open(CONFIG_FILE, 'w') as configfile:
            config.write(configfile)
        print("Inserted app name data on config.ini")
    return config


'''
    Organize data
'''

# Store data in one file
def storeDatas(file_name, raw_datas):
    f = open(file_name, mode="r", encoding="utf-8")
    lines = f.readlines()
    tr_datas = []       # tr_datas = training datas

    # 트레이닝 데이터에서 utterance만 추출
    count = 0
    for line in lines:
        if line.find('utterance') > 0:
            tr_datas.append(line)

    for tr_data in tr_datas:
        raw_data = tr_data.split('"')
        # print(raw_data)
        try:
            raw_datas.append(raw_data[1])
        except:
            raw_datas.append(raw_data[0].split('(')[1][:-2])

    f.close()
    return raw_datas

# load each training data
def loadDatas(route):
    raw_datas = []

    print('File route is ', route)
    # training data number is (0-9, a-z)
    for i in range(0, 36):
        if i < 26:
            data_number = chr(97 + i)
        else:
            data_number = str(i - 26)
        try:
            print("load file - " + route + '/resources/ko-KR/training/t-' + data_number + '.training.6t')
            storeDatas(route + '/resources/ko-KR/training/t-' + data_number + '.training.6t', raw_datas)
        except:
            print("File is none")

    print("File load done.")

    return raw_datas


def insertData(cla_datas):
    # Excel Set
    wb = Workbook()
    wsl = []
    sheet_num = 0
    # print(cla_datas)
    for sheet_name in cla_datas.keys():
        real_SN = sheet_name
        
        # Make wsl sheet name
        real_SN = real_SN.split(':continue')[0]
        real_SN = real_SN[real_SN.rfind('.')+1:].replace(':', '|')
        # print(real_SN)
        wsl.append(wb.create_sheet(real_SN[:30], sheet_num))
        # wsl.append(wb.create_sheet('temp', sheet_num)) 

        # Set Default Excel Setting
        # Will be update later...
        wsl[sheet_num]["A1"] = "Goal Name"
        wsl[sheet_num]["B1"] = "Utterance"
        wsl[sheet_num]["C1"] = "Parameter"

        # Cells merge
        wsl[sheet_num].merge_cells("A1:A2")
        wsl[sheet_num].merge_cells("B1:B2")

        # Make unique parameter list
        param_list = []
        for tr in cla_datas[sheet_name]['list']:
            params = pd.getParam(tr)
            temp = []
            for parameter in params:
                parameter = parameter[parameter.rfind('.')+1:]
                if parameter.rfind(':') == -1:
                    temp.append(parameter[:-1])
                else:
                    temp.append(parameter[:parameter.rfind(':')])
            
            set_pl = set(param_list)
            set_p = set(temp)
            for parameter in set_p - set_pl:
                param_list.append(parameter)
                # print(parameter)
        if '' in param_list:
            param_list.remove('')

        # Insert parameter names on cell
        for i in range(len(param_list)):
            wsl[sheet_num][chr(67 + i) + '2'] = param_list[i]

        except_ = 'viv.'
        except_ += config['DEFAULT']['app_name'] + '.'
        row = 3
        for tr in cla_datas[sheet_name]['list']:
            # Check config ini & Insert datas
            if config['DEFAULT']['full_goal_name'] == 'False':
                wsl[sheet_num]["A" + str(row)] = pd.getGoalName(tr).replace(except_, '')
            else:
                wsl[sheet_num]["A" + str(row)] = pd.getGoalName(tr)

            if config['DEFAULT']['full_utterance'] == 'True':
                wsl[sheet_num]["B" + str(row)] = tr
            else:
                wsl[sheet_num]["B" + str(row)] = pd.pureData(tr)
            
            # parameter pattern value
            params = pd.getParam(tr)
            cnt = 0
            param_num = 0
            for param in param_list:
                for i in range(len(params)):
                    if params[i].find(param) > 0:
                        wsl[sheet_num][chr(67 + cnt) + str(row)] = params[i]
                        param_num += 2 ** cnt
                cnt += 1
            wsl[sheet_num][chr(67 + cnt + 1) + str(row)] = param_num
            row += 1
        wsl[sheet_num].merge_cells("C1:" + chr(67 + cnt) + "1")     # parameter cell merge
        sheet_num += 1
    print("Inserting done.")
    
    # print('checking point')
    wsl.append(wb.create_sheet("summary", sheet_num))
    wsl[sheet_num]["A1"] = "Goal Name"
    wsl[sheet_num]["B1"] = "Inserted Data"
    wsl[sheet_num]["C1"] = "비고"
    summary_row = 2
    for goal in cla_datas.keys():
        wsl[sheet_num]["A" + str(summary_row)] = goal
        wsl[sheet_num]["B" + str(summary_row)] = cla_datas[goal]["count"]
        summary_row += 1

    now = datetime.datetime.now().strftime('_%Y%m%d_%H%M')
    if not os.path.isdir(SAVE_ROUTE):
        os.mkdir('./' + SAVE_ROUTE)
    wb.save(config['DEFAULT']['save_route'] + "inserted_data_" + config['DEFAULT']['app_name'] + now + ".xlsx")

    
def new_insertData(cla_datas):
    print("Write param_cla.txt")
    with open("param_cla.txt", 'w') as f:
        f.write(pj.pretty(cla_datas))
    
    # Excel Set
    wb = Workbook()
    wsl = []

    sheet_num = 0

    
    for goal_name in cla_datas.keys():
        # goal_name = goal_name.replace(':', '|')
        # print(goal_name.replace(':', '|')[:30])
        # sheet name으로 '*' 문자가 불가능 하므로 'star'로 대체
        
        sheet_name = goal_name.replace(':', '|')[:30]

        if sheet_name.find('*') >= 0:
            sheet_name = sheet_name.replace('*', 'star')

        try:
            wsl.append(wb.create_sheet(sheet_name, sheet_num))
        except:
            wsl.append(wb.create_sheet('temp', sheet_num))

        wsl[sheet_num]['B1'] = 'Goal Name:'
        wsl[sheet_num]['B2'] = 'Count:'
        wsl[sheet_num]['C1'] = goal_name
        wsl[sheet_num]['C2'] = cla_datas[goal_name]['count']

        for cell in ['B1', 'B2']:
            wsl[sheet_num][cell].alignment = Alignment(horizontal='right')
        wsl[sheet_num]['C2'].alignment = Alignment(horizontal='left')
        wsl[sheet_num]['C2'].font = Font(bold=True)

        wsl[sheet_num]['B4'] = 'Parameters'
        wsl[sheet_num]['B4'].fill = PatternFill(patternType='solid', fgColor = Color('FF9646'))
        cell_num = 'C'
        cell_cols = 0
        for param_name in cla_datas[goal_name]['param_spec']:
            # print(type(param_name))
            wsl[sheet_num][chr(ord(cell_num) + cell_cols) + '4'] = param_name
            wsl[sheet_num][chr(ord(cell_num) + cell_cols) + '4'].fill = PatternFill(patternType='solid', fgColor = Color('FF9646'))    
            cell_cols += 1
        
        # param
        param = []
        for parameter in cla_datas[goal_name].keys():
            if parameter == 'count' or parameter == 'param_spec':
                pass
            else:
                param.append(parameter)

        # param_spec 순서 재정렬
        done = False
        count = 1
        param_list = []
        while not done:
            for p in param:
                if len(p) == count:
                    param_list.append(p)

            if len(param_list) == len(param):
                done = True
            count += 1
        
        # print(param_list)
        row_count = 5
        for param in param_list:
            
            pattern_list = sorted(cla_datas[goal_name][param])
            for pattern in pattern_list:
                ad= ('B' + str(row_count) + ':' +  chr(ord('B') + cell_cols) + str(row_count))
                try:
                    wsl[sheet_num].merge_cells(ad)
                except:
                    pass
                wsl[sheet_num]['B' + str(row_count)] = '#' + pattern
                wsl[sheet_num]['B' + str(row_count)].fill = PatternFill(patternType='solid', fgColor = Color('FFBE8C'))
                row_count += 1
                data_list = sorted(cla_datas[goal_name][param][pattern])
                for data in data_list:
                    wsl[sheet_num]['A' + str(row_count)] = data                 # raw data
                    wsl[sheet_num]['B' + str(row_count)] = pd.pureData(data)    # pure data
                    # parameter 별
                    # wsl[sheet_num]['D' + str(row_count)] = str(pd.getParam(data))
                    for data_param in pd.getParam(data):
                        if not data_param == '':
                            # data_param의 'v:' 삭제
                            data_param = data_param.replace('v:viv.', '')
                            data_param = data_param.split(':')
                            # print(data_param)
                            ck = data_param[0][data_param[0].find('.') + 1:]
                            res = data_param[0][1:data_param[0].find(')')]
                            if len(data_param) > 1:
                                res += '(' + data_param[1][:-1] + ')'
                            else:
                                ck = ck[:-1]
                            # print(ck)

                            try:
                                index = list(cla_datas[goal_name]['param_spec']).index(ck)        
                                wsl[sheet_num][chr(ord('C') + index) + str(row_count)] = res
                            except: pass
                    row_count += 1
                row_count += 1
            
        
        # Next sheet
        sheet_num += 1


    # Excel file save
    print("Inserting done.")
    now = datetime.datetime.now().strftime('_%Y%m%d_%H%M')
    if not os.path.isdir(SAVE_ROUTE):
        os.mkdir('./' + SAVE_ROUTE)
    
    # 저장 시간이 겹쳐서 에러가 날 경우 
    try:
        wb.save(config['DEFAULT']['save_route'] + "inserted_data_" + config['DEFAULT']['app_name'] + now + ".xlsx")
    except:
        wb.save(config['DEFAULT']['save_route'] + "inserted_data_" + config['DEFAULT']['app_name'] + now + "_.xlsx")

    print("File saved.")

def makePatternView(cla_datas):
    pattern = {}

    # goal_name 오름차순
    goal_name_list = sorted(cla_datas.keys())
    
    for goal_name in goal_name_list:
    
        param_list = sorted(list(cla_datas[goal_name]['param_count'].keys()))
        # print('param_list: ' + str(param_list))

        temp_goal_name = goal_name.replace("viv." + config['DEFAULT']['app_name'] + '.', '') 
        pattern[temp_goal_name] = {}
        pattern[temp_goal_name]['count'] = cla_datas[goal_name]['count']
        pattern[temp_goal_name]['param_spec'] = cla_datas[goal_name]['param_spec']
                            
        for param in param_list:
            pattern[temp_goal_name][param] = {}
        
        cla_datas[goal_name]['param_count'][param] = sorted(cla_datas[goal_name]['param_count'][param])
        
        for key in cla_datas[goal_name]['param_count'].keys():
            for data in cla_datas[goal_name]['param_count'][key]:
                # print(key + ' : ' + data)
                new_data = returnPattern(data)

                if not new_data in pattern[temp_goal_name][key]:
                    pattern[temp_goal_name][key][new_data] = []
                
                pattern[temp_goal_name][key][new_data].append(data)
        

    return pattern

def returnPattern(data):
    # pattern config setting
    patterns = []
    last_pattern = config['pattern']['pp'].split(',')
    # real_last_patterns = []
    for pat in config['pattern'].keys():
        if pat.find('pa_') == 0:
            patterns.append(config['pattern'][pat].split(','))

    # value 값으로 변경
    while data.find('(') >= 0:
        start = data.find('(')
        end = data[start:].find(']') + start
        old = data[start:end + 1]

        n_start = data[:end].rfind('.') + 1

        if data[n_start:end].find(':') > 0:
            end = data[n_start:end].find(':') + n_start
            

        new = data[n_start:end]
        data = data.replace(old, '[' + new + ']')

    data_pattern = ''
    for d in data.split():

        # goal 삭제
        if not d.find('g:') == -1:
            continue

        # parameter 개수 확인
        # 파라미터 조합 별 나눠야 함
        
        # print(d)
        d_len = len(d)
        for pattern in last_pattern:
            # 각 어절 마지막이 pattern으로 끝나는 경우 제거
            if d_len == len(pattern) + d.find(pattern):
                d = d.replace(pattern, '')
        
        data_pattern += d + ' '
        
    data_pattern = ' '.join(data_pattern.split())
    
    # 사용자 추가 발화로 수정
    for pattern in patterns:
        for index in range(1, len(pattern)):
            data_pattern = data_pattern.replace(pattern[index], pattern[0])
    
    return data_pattern
    
'''
    Main process
'''

def main():
    if not os.path.isfile(CONFIG_FILE):
        makeConfig()
    config = readConfig()

    file_route = config['DEFAULT']['file_route'] + config['DEFAULT']['app_name']

    raw_datas = loadDatas(file_route)
    cla_datas = pd.classifyDatas(raw_datas)     # cla_datas: classfied datas
    
    if config['DEFAULT']['classify_dictionary'] == 'True':
        pd.makeEasyToReadable(cla_datas)

    if config['DEFAULT']['pattern_view'] == 'True':
        new_cladatas = makePatternView(cla_datas)
        new_insertData(new_cladatas)
    else:
        insertData(cla_datas)
    
if __name__ == "__main__":
    main()